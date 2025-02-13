import openpyxl


class HomePageData:
    test_HomePage_data = [{"firstname": "Rahul", "lastname": "Shetty", "gender": "Male"},
                          {"firstname": "Anshika", "lastname": "Shetty", "gender": "Female"},
                          {"firstname": "Tom", "lastname": "Cruise", "gender": "Male"}]

    @staticmethod
    def get_test_data(test_case_name):
        dictionary = {}
        book = openpyxl.load_workbook(r"C:\Users\grzeg\PycharmProjects\PythonSelfFramework\test_data\PythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    dictionary[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [dictionary]
