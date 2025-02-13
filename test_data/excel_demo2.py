import openpyxl

book = openpyxl.load_workbook(r"C:\Users\grzeg\Documents\SeleniumTestFiles\download.xlsx")
sheet = book.active
dictionary = {}


edit_row = 0
edit_column = 0

j = 1

for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=j).value == "Apple":
        edit_row = i
    else:
        j = j + 1


print(edit_row)
