from time import sleep

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions
import openpyxl


def update_excel_data(filepath, search_term, column_name, new_value):
    book = openpyxl.load_workbook(filepath)
    sheet = book.active
    dictionary = {}

    for i in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=i).value == column_name:
            dictionary["column"] = i

    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            if sheet.cell(row=i, column=j).value == search_term:
                dictionary["row"] = i

    sheet.cell(row=dictionary["row"], column=dictionary["column"]).value = new_value

    book.save(file_path)


driver = webdriver.Chrome()
driver.implicitly_wait(5)
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")

file_path = r"C:\Users\grzeg\Downloads\download.xlsx"
fruit_name = "Apple"
new_value = "999"

driver.find_element(By.ID, "downloadButton").click()

sleep(2)

# edit excel file

update_excel_data(file_path, fruit_name, "price", "999")

file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
file_input.send_keys(file_path)
toast_locator = (By.CSS_SELECTOR, ".Toastify__toast-body div:nth-child(2)")
WebDriverWait(driver, 5).until(expected_conditions.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)

price_column = driver.find_element(By.XPATH, "//div[text()='Price']").get_attribute("data-column-id")
actual_price = driver.find_element(By.XPATH,
                                   "//div[text()='" + fruit_name + "']/parent::div/parent::div/div[@id='cell-" + price_column + "-undefined']/div").text
assert actual_price == new_value
